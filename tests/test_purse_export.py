import io
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

from tests.conftest import create_org_and_group, find_redis_token, onboard_group_admin


async def _register_and_login_group_admin(client, email="export-rep@example.com"):
    await client.post(
        "/auth/register",
        json={
            "email": email,
            "password": "P@ssword123",
            "first_name": "Tayo",
            "last_name": "Rep",
            "role": "group_admin",
        },
    )
    verify_token = await find_redis_token("verify_email")
    await client.post("/auth/verify-email", json={"email": email, "token": verify_token})
    login = await client.post("/auth/login", json={"email": email, "password": "P@ssword123"})
    return login.json()["data"]["access_token"]


async def _register_and_login_member(client, token, email, first_name="Ada", last_name="Lovelace"):
    await client.post(
        f"/members/join/{token}",
        json={"email": email, "password": "P@ssword123", "first_name": first_name, "last_name": last_name},
    )
    verify_token = await find_redis_token("verify_email")
    await client.post("/auth/verify-email", json={"email": email, "token": verify_token})
    login = await client.post("/auth/login", json={"email": email, "password": "P@ssword123"})
    return login.json()["data"]["access_token"]


def _future_deadline(days=7) -> str:
    return (datetime.now(timezone.utc) + timedelta(days=days)).isoformat()


async def _setup_purse_with_member(client, db_session, email="export-rep@example.com", member_email="export-member@example.com"):
    org, _existing_group = await create_org_and_group(db_session)
    admin_token = await _register_and_login_group_admin(client, email=email)
    admin_headers = {"Authorization": f"Bearer {admin_token}"}
    group = await onboard_group_admin(client, db_session, org, admin_headers)

    invite = await client.post(
        f"/group-admins/invite-links?group_id={group.id}", json={"expires_in_days": 7}, headers=admin_headers
    )
    token = invite.json()["data"]["token"]
    await _register_and_login_member(client, token, member_email)

    create = await client.post(
        "/purses",
        json={
            "group_id": str(group.id),
            "title": "Excursion Fee",
            "amount": "1000.00",
            "deadline": _future_deadline(),
            "enroll_mode": "snapshot",
        },
        headers=admin_headers,
    )
    purse_id = create.json()["data"]["id"]
    return org, group, admin_headers, purse_id


async def test_export_xlsx_has_correct_headers_and_matches_contributions(client, db_session):
    org, group, admin_headers, purse_id = await _setup_purse_with_member(client, db_session)

    resp = await client.get(f"/purses/{purse_id}/export?format=xlsx", headers=admin_headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    disposition = resp.headers["content-disposition"]
    assert disposition.startswith("attachment;")
    assert "excursion-fee-" in disposition
    assert disposition.endswith('.xlsx"')

    wb = load_workbook(io.BytesIO(resp.content))
    ws = wb.active
    assert ws["A1"].value == "Excursion Fee"

    rows = list(ws.iter_rows(values_only=True))
    header_index = next(i for i, row in enumerate(rows) if row and row[0] == "Member")
    assert rows[header_index] == (
        "Member",
        "Member ID",
        "Status",
        "Amount expected",
        "Amount received",
        "Paid at",
    )
    data_row = rows[header_index + 1]
    assert data_row[0] == "Ada Lovelace"
    assert data_row[3] == "1000.00"
    assert data_row[4] == "0.00"


async def test_export_pdf_returns_pdf_content(client, db_session):
    org, group, admin_headers, purse_id = await _setup_purse_with_member(client, db_session)

    resp = await client.get(f"/purses/{purse_id}/export?format=pdf", headers=admin_headers)
    assert resp.status_code == 200, resp.text
    assert resp.headers["content-type"] == "application/pdf"
    disposition = resp.headers["content-disposition"]
    assert "excursion-fee-" in disposition
    assert disposition.endswith('.pdf"')
    assert resp.content.startswith(b"%PDF")


async def test_export_invalid_format_rejected(client, db_session):
    org, group, admin_headers, purse_id = await _setup_purse_with_member(client, db_session)

    resp = await client.get(f"/purses/{purse_id}/export?format=csv", headers=admin_headers)
    assert resp.status_code == 422


async def test_export_requires_admin_of_the_purses_own_group(client, db_session):
    org, group, admin_headers, purse_id = await _setup_purse_with_member(
        client, db_session, email="export-owner@example.com", member_email="export-owner-member@example.com"
    )

    other_org, _existing_other_group = await create_org_and_group(
        db_session, org_name="Other Uni", org_short_code="OU9", group_name="Other Dept", group_short_code="OD9"
    )
    other_admin_token = await _register_and_login_group_admin(client, email="export-other-rep@example.com")
    other_headers = {"Authorization": f"Bearer {other_admin_token}"}
    await onboard_group_admin(client, db_session, other_org, other_headers, group_name="Other Group")

    resp = await client.get(f"/purses/{purse_id}/export?format=xlsx", headers=other_headers)
    assert resp.status_code == 403


async def test_export_requires_authentication(client, db_session):
    org, group, admin_headers, purse_id = await _setup_purse_with_member(client, db_session)

    resp = await client.get(f"/purses/{purse_id}/export?format=xlsx")
    assert resp.status_code == 401
