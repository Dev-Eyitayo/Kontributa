import io
import re
from datetime import datetime, timezone
from decimal import Decimal
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
from reportlab.platypus import Table as PdfTable
from reportlab.platypus import TableStyle

from app.modules.contributions.service import ContributionRow, compute_display_status
from app.modules.purses.models import Purse

ExportFormat = Literal["xlsx", "pdf"]

_HEADER = ["Member", "Member ID", "Status", "Amount expected", "Amount received", "Paid at"]
_CONTENT_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def _report_rows(contributions: list[ContributionRow], purse_status: str) -> list[list[str]]:
    return [
        [
            # An admin's own contribution is just another row in this same
            # report -- no separate section -- with a minimal inline label
            # so it's still distinguishable at a glance in a flat export.
            f"{row.name} (Admin)" if row.owner_type == "admin" else row.name,
            row.member_id_number or "",
            compute_display_status(row.contribution.status.value, purse_status),
            str(row.contribution.amount_expected),
            str(row.contribution.amount_received),
            row.contribution.paid_at.isoformat() if row.contribution.paid_at else "",
        ]
        for row in contributions
    ]


def export_filename(purse_title: str, fmt: ExportFormat) -> str:
    # ASCII-safe slug -- a purse title is free text and could contain
    # anything (slashes, quotes, emoji), none of which belong in a
    # Content-Disposition filename.
    slug = re.sub(r"[^a-zA-Z0-9]+", "-", purse_title).strip("-").lower() or "purse"
    date_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    return f"{slug}-{date_str}.{fmt}"


def content_type_for(fmt: ExportFormat) -> str:
    return _CONTENT_TYPES[fmt]


def build_xlsx(purse: Purse, contributions: list[ContributionRow], total_collected: Decimal) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Contributions"

    ws.append([purse.title])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Deadline: {purse.deadline.isoformat()}"])
    ws.append([f"Total collected: {total_collected}"])
    ws.append([f"Generated at: {datetime.now(timezone.utc).isoformat()}"])
    ws.append([])

    ws.append(_HEADER)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    for row in _report_rows(contributions, purse.status.value):
        ws.append(row)

    for column_cells in ws.columns:
        values = [str(c.value) for c in column_cells if c.value is not None]
        width = max((len(v) for v in values), default=10)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_pdf(purse: Purse, contributions: list[ContributionRow], total_collected: Decimal) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()

    elements = [
        Paragraph(purse.title, styles["Title"]),
        Paragraph(f"Deadline: {purse.deadline.isoformat()}", styles["Normal"]),
        Paragraph(f"Total collected: {total_collected}", styles["Normal"]),
        Paragraph(f"Generated at: {datetime.now(timezone.utc).isoformat()}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    data = [_HEADER, *_report_rows(contributions, purse.status.value)]
    table = PdfTable(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f3d2e")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buffer.getvalue()


def build_export(
    fmt: ExportFormat, purse: Purse, contributions: list[ContributionRow], total_collected: Decimal
) -> bytes:
    if fmt == "xlsx":
        return build_xlsx(purse, contributions, total_collected)
    return build_pdf(purse, contributions, total_collected)
